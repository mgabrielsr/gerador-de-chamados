from flask import Flask, render_template, request, redirect, url_for
import pandas as pd
import logging
from notifier import send_email, send_sms
from dotenv import load_dotenv
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# Carregar variáveis do arquivo .env
load_dotenv()

# Obter o número de telefone para enviar SMS do arquivo .env
SMS_PHONE_NUMBER = os.getenv('SMS_PHONE_NUMBER')

app = Flask(__name__, template_folder='../frontend/templates', static_folder='../frontend/static')

# Configurar o log
logging.basicConfig(level=logging.DEBUG)

@app.route('/')
def index():
    return render_template('form_chamado.html')

@app.route('/abrir_chamado', methods=['POST'])
def abrir_chamado():
    try:
        data = request.form['data']
        horario = request.form['horario']
        numero_chamado = request.form['numero_chamado']
        endereco = request.form['endereco']
        numero = request.form['numero']
        cep = request.form['cep']
        cidade = request.form['cidade']
        nome_empresa = request.form['nome_empresa']
        contato_empresa = request.form['contato_empresa']
        responsavel = request.form['responsavel']
        defeito = request.form['defeito']
        servico = request.form['servico']
        modelo = request.form['modelo']
        numero_serie = request.form['numero_serie']
        status = request.form['status']
        observacoes = request.form.get('observacoes', '') 

        # Log dos dados recebidos
        logging.debug(f"Data: {data}")
        logging.debug(f"Horário: {horario}")
        logging.debug(f"Número do Chamado: {numero_chamado}")
        logging.debug(f"Endereço: {endereco}")
        logging.debug(f"Número: {numero}")
        logging.debug(f"CEP: {cep}")
        logging.debug(f"Cidade: {cidade}")
        logging.debug(f"Nome da Empresa: {nome_empresa}")
        logging.debug(f"Contato da Empresa: {contato_empresa}")
        logging.debug(f"Responsável: {responsavel}")
        logging.debug(f"Defeito: {defeito}")
        logging.debug(f"Serviço: {servico}")
        logging.debug(f"Modelo: {modelo}")
        logging.debug(f"Número de Série: {numero_serie}")
        logging.debug(f"Status: {status}")
        logging.debug(f"Observações: {observacoes}")

    except KeyError as e:
        return f"Missing form field: {e}", 400

    # Carregar dados do arquivo Excel existente
    excel_file = 'agenda_chamados.xlsx'
    try:
        df = pd.read_excel(excel_file, engine='openpyxl')
    except FileNotFoundError:
        df = pd.DataFrame(columns=[
            'data', 'horario', 'numero_chamado', 'endereco', 'numero', 'cep', 'cidade', 
            'nome_empresa', 'contato_empresa', 'responsavel', 'defeito', 'servico', 
            'modelo', 'numero_serie', 'status', 'observacoes'
        ])
        
    # Adicionar novo chamado
    novo_chamado = pd.DataFrame([{
        'data': data,
        'horario': horario,
        'numero_chamado': numero_chamado,
        'endereco': endereco,
        'numero': numero,
        'cep': cep,
        'cidade': cidade,
        'nome_empresa': nome_empresa,
        'contato_empresa': contato_empresa,
        'responsavel': responsavel,
        'defeito': defeito,
        'servico': servico,
        'modelo': modelo,
        'numero_serie': numero_serie,
        'status': status,
        'observacoes': observacoes
    }])
    
    df = pd.concat([df, novo_chamado], ignore_index=True)

    # # Verificar o DataFrame para depuração
    # print("DataFrame atualizado:")
    # print(df)

    # Salvar os dados atualizados de volta no arquivo Excel
    df.to_excel(excel_file, index=False, engine='openpyxl')
    
    # Recarregar o arquivo Excel para aplicar formatação
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    # Aplicar alinhamento à direita para as colunas E e I
    for row in range(2, sheet.max_row + 1):  # Começa da linha 2 para ignorar o cabeçalho
        sheet[f'E{row}'].alignment = Alignment(horizontal='right')
        sheet[f'I{row}'].alignment = Alignment(horizontal='right')

    # Definir larguras das colunas
    column_widths = {
        'A': 10,
        'B': 9,
        'C': 20,
        'D': 40,
        'E': 9,
        'F': 9,
        'G': 15,
        'H': 18,
        'I': 20,
        'J': 30,
        'K': 30,
        'L': 30,
        'M': 14,
        'N': 18,
        'O': 15,
        'P': 14
    }

    for column, width in column_widths.items():
        sheet.column_dimensions[column].width = width
        
    # Ajustar a altura da linha 1
    sheet.row_dimensions[1].height = 25

    # Definir negrito e centralizar o texto na linha 1
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Adicionar a validação de dados (lista suspensa) na célula O2 até a última linha usada
    data_validation = DataValidation(
        type="list",
        formula1='"Aberto,Em Andamento,Finalizado"',
        showDropDown=True
    )
    sheet.add_data_validation(data_validation)

    # Aplicar a validação na coluna O
    for row in range(2, sheet.max_row + 1):
        cell = sheet[f'O{row}']
        data_validation.add(cell)

    # Definir a formatação condicional para as cores
    colors = {
        'Aberto': '00FF00',          # Verde
        'Em Andamento': 'FFFF00'     # Amarelo
    }

    for status, color in colors.items():
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        rule = CellIsRule(
            operator='equal',
            formula=[f'"{status}"'],
            fill=fill
        )
        # Adicionando a formatação condicional à coluna O
        sheet.conditional_formatting.add(f'O2:O{sheet.max_row}', rule)

        # Ajustar a largura da coluna O
        sheet.column_dimensions['O'].width = 15

    # Adicionar bordas a todas as células
    border = Border(
        left=Side(border_style='thick'),
        right=Side(border_style='thick'),
        top=Side(border_style='thick'),
        bottom=Side(border_style='thick')
    )

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border

    # Remover linhas de grade
    sheet.sheet_view.showGridLines = False

    # Salvar as alterações na planilha
    workbook.save(excel_file)
    
    # Conteúdo da notificação
    subject = f"Novo Chamado Aberto - {numero_chamado}"
    body = f"""
    Prezado(a) (Nome do Técnico),

    Gostaríamos de informar que um novo chamado foi criado no sistema. Abaixo estão os detalhes do chamado:

    - Número do Chamado: {numero_chamado}
    - Data: {data}
    - Horário: {horario}
    - Endereço: {endereco}
    - Número: {numero}
    - CEP: {cep}
    - Cidade: {cidade}
    - Nome da Empresa: {nome_empresa}
    - Contato da Empresa: {contato_empresa}
    - Responsável: {responsavel}
    - Defeito: {defeito}
    - Serviço: {servico}
    - Modelo: {modelo}
    - Número de Série: {numero_serie}
    - Status: {status}
    - Observações: {observacoes}

    Por favor, revise os detalhes e tome as providências necessárias. Para mais informações, consulte a planilha de chamados atualizada.

    Caso tenha alguma dúvida ou precise de mais informações, não hesite em entrar em contato.

    Atenciosamente,

    [Seu Nome]  
    [Seu Cargo]  
    [Seu Contato]  
    [Seu E-mail]  
    [Nome da Empresa]
    """
    # Caminho do arquivo Excel
    excel_file = 'agenda_chamados.xlsx'

    # Enviar notificação por e-mail
    send_email(subject, body, attachment_path=excel_file)

    # Conteúdo da notificação por SMS
    sms_body = (
        f"Novo Chamado Recebido!\n\n"
        f"Número do Chamado: {numero_chamado}\n"
        f"Data e Horário: {data} às {horario}\n"
        f"Empresa: {nome_empresa}\n"
        f"Endereço: {endereco}, {numero}, {cidade} - {cep}\n"
        f"Defeito Reportado: {defeito}\n"
        f"Serviço Solicitado: {servico}\n"
        f"Status Atual: {status}\n\n"
        f"Consulte a planilha de chamados para mais detalhes e atualizações.\n\n"
        f"Obrigado!"
    )

    # Número para o qual o SMS será enviado
    phone_number = SMS_PHONE_NUMBER 

    # Enviar notificação por SMS
    send_sms(phone_number, sms_body)

    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)
